import csv
import io
import json
from datetime import datetime, timezone
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.services.vehicle_export import (
    EXPORT_COLUMNS,
    ExportFormat,
    _safe_filename,
    _vehicle_data_to_record,
    build_vehicle_export,
)


@pytest.fixture
def vehicle_rows():
    """
    Lightweight stand-ins for VehicleData SQLAlchemy objects.

    The export service only reads attributes from each object,
    so these tests do not require a database.
    """
    return [
        SimpleNamespace(
            vehicle_id="car-1",
            timestamp=datetime(
                2026,
                7,
                21,
                10,
                30,
                tzinfo=timezone.utc,
            ),
            speed=45.5,
            odometer=12000.25,
            soc=82.0,
            elevation=35.2,
            shift_state="D",
        ),
        SimpleNamespace(
            vehicle_id="car-1",
            timestamp=datetime(
                2026,
                7,
                21,
                10,
                31,
                tzinfo=timezone.utc,
            ),
            speed=48.0,
            odometer=12001.0,
            soc=81.5,
            elevation=36.0,
            shift_state="D",
        ),
    ]


def test_safe_filename_replaces_unsafe_characters():
    assert _safe_filename("vehicle/one:dangerous") == (
        "vehicle_one_dangerous"
    )


def test_safe_filename_returns_default_when_empty():
    assert _safe_filename("") == "vehicle"


def test_vehicle_data_to_record(vehicle_rows):
    record = _vehicle_data_to_record(vehicle_rows[0])

    assert record == {
        "vehicle_id": "car-1",
        "timestamp": "2026-07-21T10:30:00+00:00",
        "speed": 45.5,
        "odometer": 12000.25,
        "soc": 82.0,
        "elevation": 35.2,
        "shift_state": "D",
    }


def test_build_json_export(vehicle_rows):
    exported = build_vehicle_export(
        rows=vehicle_rows,
        vehicle_id="car-1",
        export_format=ExportFormat.JSON,
    )

    assert exported.filename == "car-1_vehicle_data.json"
    assert exported.media_type == "application/json"

    content = json.loads(exported.content.decode("utf-8"))

    assert len(content) == 2
    assert content[0]["vehicle_id"] == "car-1"
    assert content[0]["timestamp"] == "2026-07-21T10:30:00+00:00"
    assert content[0]["speed"] == 45.5
    assert content[1]["speed"] == 48.0


def test_build_csv_export(vehicle_rows):
    exported = build_vehicle_export(
        rows=vehicle_rows,
        vehicle_id="car-1",
        export_format=ExportFormat.CSV,
    )

    assert exported.filename == "car-1_vehicle_data.csv"
    assert exported.media_type == "text/csv; charset=utf-8"

    # utf-8-sig removes the BOM added by _build_csv.
    text = exported.content.decode("utf-8-sig")
    records = list(csv.DictReader(io.StringIO(text)))

    assert records[0] == {
        "vehicle_id": "car-1",
        "timestamp": "2026-07-21T10:30:00+00:00",
        "speed": "45.5",
        "odometer": "12000.25",
        "soc": "82.0",
        "elevation": "35.2",
        "shift_state": "D",
    }

    assert records[1]["speed"] == "48.0"
    assert tuple(records[0].keys()) == EXPORT_COLUMNS


def test_build_excel_export(vehicle_rows):
    exported = build_vehicle_export(
        rows=vehicle_rows,
        vehicle_id="car-1",
        export_format=ExportFormat.XLSX,
    )

    assert exported.filename == "car-1_vehicle_data.xlsx"
    assert exported.media_type == (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )

    workbook = load_workbook(
        filename=io.BytesIO(exported.content),
        read_only=True,
        data_only=True,
    )

    worksheet = workbook["Vehicle data"]
    spreadsheet_rows = list(worksheet.iter_rows(values_only=True))

    assert spreadsheet_rows[0] == EXPORT_COLUMNS

    assert spreadsheet_rows[1] == (
        "car-1",
        "2026-07-21T10:30:00+00:00",
        45.5,
        12000.25,
        82.0,
        35.2,
        "D",
    )

    assert spreadsheet_rows[2][0] == "car-1"
    assert spreadsheet_rows[2][2] == 48.0

    workbook.close()


@pytest.mark.parametrize(
    ("export_format", "extension"),
    [
        (ExportFormat.JSON, "json"),
        (ExportFormat.CSV, "csv"),
        (ExportFormat.XLSX, "xlsx"),
    ],
)
def test_build_export_sanitises_vehicle_id(
    vehicle_rows,
    export_format,
    extension,
):
    exported = build_vehicle_export(
        rows=vehicle_rows,
        vehicle_id="../../car one",
        export_format=export_format,
    )

    assert exported.filename == (
        ".._.._car_one_vehicle_data."
        f"{extension}"
    )

    assert "/" not in exported.filename
    assert " " not in exported.filename


@pytest.mark.parametrize(
    "export_format",
    [
        ExportFormat.JSON,
        ExportFormat.CSV,
        ExportFormat.XLSX,
    ],
)
def test_build_export_handles_empty_rows(export_format):
    exported = build_vehicle_export(
        rows=[],
        vehicle_id="empty-car",
        export_format=export_format,
    )

    assert exported.content
    assert exported.filename.startswith("empty-car_vehicle_data")
